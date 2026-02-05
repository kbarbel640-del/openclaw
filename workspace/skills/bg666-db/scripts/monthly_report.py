#!/usr/bin/env python3
"""
BG666 月度報表生成器
從 DB 拉數據，生成跟現有 Excel 格式一致的報表

來源表：
- channel_data_statistics: 注冊/活躍/ip/充值/提現/首充/充值人數
- channel_game_statistics: 總投注

口徑（2026-01-31 確認，全部 ✅ 對齊）：
- 注冊 = SUM(register_number)
- 活躍 = SUM(active_number)
- ip人數 = SUM(ip_number)
- 充值 = SUM(recharge_amount)
- 提現 = SUM(withdraw_amount)
- 充提差 = 充值 - 提現
- 總投注 = SUM(bet_amount) from channel_game_statistics
- 首充 = SUM(first_recharge_number)
- 充值人數 = SUM(recharge_number)
- 人均充值 = 充值 / 充值人數
"""

import argparse
import calendar
import json
import os
import sys
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent

DB_CONFIG = {
    'host': 'bg666-market-readonly.czsks2mguhd5.ap-south-1.rds.amazonaws.com',
    'port': 3306,
    'user': 'market',
    'password': 'hBVoVVm&)aZtW0t6',
    'database': 'ry-cloud',
    'connect_timeout': 10,
    'charset': 'utf8mb4'
}

def get_db_connection():
    import pymysql
    return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)

def query_monthly_data(year: int, month: int):
    """Pull all metrics for a given month."""
    conn = get_db_connection()
    days_in_month = calendar.monthrange(year, month)[1]
    start = f"{year}-{month:02d}-01"
    end = f"{year}-{month:02d}-{days_in_month}"
    
    try:
        with conn.cursor() as cur:
            # Main metrics from channel_data_statistics
            cur.execute("""
                SELECT statistics_day,
                    SUM(register_number) AS reg,
                    SUM(active_number) AS active,
                    SUM(ip_number) AS ip,
                    SUM(recharge_amount) AS recharge,
                    SUM(withdraw_amount) AS withdraw,
                    SUM(recharge_amount) - SUM(withdraw_amount) AS diff,
                    SUM(first_recharge_number) AS first_charge,
                    SUM(recharge_number) AS charge_users
                FROM channel_data_statistics
                WHERE statistics_day >= %s AND statistics_day <= %s
                GROUP BY statistics_day
                ORDER BY statistics_day
            """, (start, end))
            main_data = {str(r['statistics_day']): r for r in cur.fetchall()}
            
            # Bet from channel_game_statistics
            cur.execute("""
                SELECT statistics_day, SUM(bet_amount) AS total_bet
                FROM channel_game_statistics
                WHERE statistics_day >= %s AND statistics_day <= %s
                GROUP BY statistics_day
                ORDER BY statistics_day
            """, (start, end))
            bet_data = {str(r['statistics_day']): r['total_bet'] for r in cur.fetchall()}
    finally:
        conn.close()
    
    # Merge
    result = {}
    for day_num in range(1, days_in_month + 1):
        d = f"{year}-{month:02d}-{day_num:02d}"
        m = main_data.get(d, {})
        recharge = float(m.get('recharge', 0) or 0)
        charge_users = int(m.get('charge_users', 0) or 0)
        result[d] = {
            '注册人数': int(m.get('reg', 0) or 0),
            '活跃人数': int(m.get('active', 0) or 0),
            'ip人数': int(m.get('ip', 0) or 0),
            '充值': recharge,
            '充提差': float(m.get('diff', 0) or 0),
            '总投注': float(bet_data.get(d, 0) or 0),
            '首充人数': int(m.get('first_charge', 0) or 0),
            '充值人数': charge_users,
            '人均充值': round(recharge / charge_users, 2) if charge_users > 0 else None,
        }
    
    return result, days_in_month

def generate_excel(year: int, month: int, output_path: str):
    """Generate Excel in the same format as existing monthly reports."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, numbers
    
    data, days_in_month = query_monthly_data(year, month)
    month_name = f"{month}月"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month_name
    
    # Row labels (same order as existing Excel)
    metrics = ['注册人数', '活跃人数', 'ip人数', '充值', '充提差', '总投注', '首充人数', '充值人数', '人均充值']
    
    # Header row
    ws.cell(row=1, column=1, value='Unnamed: 0')
    ws.cell(row=1, column=2, value='日均')
    for day in range(1, days_in_month + 1):
        ws.cell(row=1, column=day + 2, value=f'{month}月{day}')
    
    # Data rows
    for row_idx, metric in enumerate(metrics, start=2):
        ws.cell(row=row_idx, column=1, value=metric)
        
        # Collect values for daily average
        values = []
        for day in range(1, days_in_month + 1):
            d = f"{year}-{month:02d}-{day:02d}"
            val = data[d].get(metric)
            col = day + 2
            if val is not None and val != 0:
                ws.cell(row=row_idx, column=col, value=val)
                values.append(val)
            elif val == 0:
                # Check if date is in the future (no data yet)
                if date(year, month, day) <= date.today():
                    ws.cell(row=row_idx, column=col, value=0)
                    values.append(0)
        
        # Daily average
        if values:
            avg = sum(values) / len(values)
            ws.cell(row=row_idx, column=2, value=round(avg, 2) if metric in ['充值', '充提差', '总投注', '人均充值'] else round(avg, 1))
    
    # Format: set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    for day in range(1, days_in_month + 1):
        col_letter = openpyxl.utils.get_column_letter(day + 2)
        ws.column_dimensions[col_letter].width = 14
    
    # Bold header row
    header_font = Font(bold=True)
    for col in range(1, days_in_month + 3):
        ws.cell(row=1, column=col).font = header_font
    
    # Bold metric names
    for row in range(2, len(metrics) + 2):
        ws.cell(row=row, column=1).font = header_font
    
    wb.save(output_path)
    return output_path

def main():
    parser = argparse.ArgumentParser(description='BG666 Monthly Report Generator')
    parser.add_argument('--year', type=int, default=datetime.now().year)
    parser.add_argument('--month', type=int, default=datetime.now().month)
    parser.add_argument('--output', type=str, default=None)
    args = parser.parse_args()
    
    if args.output is None:
        args.output = str(SCRIPT_DIR.parent.parent.parent / 'output' / f'bg666_{args.year}_{args.month:02d}_report.xlsx')
    
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    path = generate_excel(args.year, args.month, args.output)
    print(f"✅ Report generated: {path}")
    print(f"   Period: {args.year}-{args.month:02d}")

if __name__ == '__main__':
    main()
