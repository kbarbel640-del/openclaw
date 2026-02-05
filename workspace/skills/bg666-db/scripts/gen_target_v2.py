#!/usr/bin/env python3
"""
BG666 2-4月 周预期进度表（v2）
基于 12月+1月 DB 实际数据推算，全部 9 个已对齐指标
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path
from datetime import date, timedelta
import calendar

OUTPUT = Path(__file__).parent.parent.parent.parent / "output" / "bg666_2026_feb_target_v2.xlsx"

# === 基准数据（DB 实际，12/1~1/30，61天）===
BASE_DAYS = 61
BASE = {
    "注册人数":   {"total": 63293,      "daily": 1037.59},
    "活跃人数":   {"total": 2458053,    "daily": 40295.95},
    "ip人数":     {"total": 3096137,    "daily": 50756.34},
    "充值":       {"total": 249050189,  "daily": 4082789.98},
    "提现":       {"total": 216238980,  "daily": 3544901.31},
    "充提差":     {"total": 32811209,   "daily": 537888.67},
    "总投注":     {"total": 2069769048.72, "daily": 33930640.14},
    "首充人数":   {"total": 34049,      "daily": 558.18},
    "充值人数":   {"total": 392084,     "daily": 6427.61},
    "人均充值":   {"total": None,       "daily": 635.13},  # 4082789.98 / 6427.61
}

# 增长假设
GROWTH = {"2月": 0.05, "3月": 0.05, "4月": 0.05}

# 周结算日（周日）
def get_weekly_sundays(year, month):
    """Get all Sundays + month end for a given month."""
    days_in = calendar.monthrange(year, month)[1]
    first = date(year, month, 1)
    last = date(year, month, days_in)
    
    sundays = []
    d = first
    while d <= last:
        if d.weekday() == 6:  # Sunday
            sundays.append(d)
        d += timedelta(days=1)
    
    # Always include month end if not already a Sunday
    if last not in sundays:
        sundays.append(last)
    
    return sorted(sundays)

def main():
    wb = openpyxl.Workbook()
    
    hfont = Font(bold=True, size=11)
    tfont = Font(bold=True, size=14)
    sfont = Font(bold=True, size=11, color="666666")
    blue = PatternFill(start_color="D6EAF8", fill_type="solid")
    green = PatternFill(start_color="D5F5E3", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    money = "#,##0"
    money2 = "#,##0.00"
    pct = "0.0%"
    
    def borders(ws, r1, r2, c1, c2):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).border = border

    # === Sheet 1: 总览 ===
    ws = wb.active
    ws.title = "总览"
    
    ws.merge_cells("A1:F1")
    ws["A1"] = "BG666 2026 2-4月 周预期进度表"
    ws["A1"].font = tfont
    ws["A2"] = "基于 DB 实际数据推算（channel_data_statistics + channel_game_statistics）"
    ws["A2"].font = sfont
    ws["A3"] = f"基准期：2025-12-01 ~ 2026-01-30（{BASE_DAYS}天）"
    ws["A4"] = "口径：自然日 | 周结算：周日 | 全部指标已与 Excel 对齐 ✅"
    
    # 基准日均
    r = 6
    ws.cell(r, 1, "基准日均指标").font = hfont
    r += 1
    headers = ["指标", "61天总量", "日均", "DB 字段", "口径说明"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h)
        c.font = hfont
        c.fill = blue
    hr = r
    r += 1
    
    sources = {
        "注册人数": "register_number",
        "活跃人数": "active_number", 
        "ip人数": "ip_number",
        "充值": "recharge_amount",
        "提现": "withdraw_amount",
        "充提差": "recharge - withdraw",
        "总投注": "bet_amount (game表)",
        "首充人数": "first_recharge_number",
        "充值人数": "recharge_number",
        "人均充值": "充值 / 充值人数",
    }
    
    notes = {
        "注册人数": "当日新注册玩家数。来源：channel_data_statistics.register_number，按 statistics_day 汇总所有渠道。与后台「注册人数」完全对齐。",
        "活跃人数": "当日至少登录一次的独立玩家数。来源：channel_data_statistics.active_number。等价于 player_logininfor 表 COUNT(DISTINCT player_id)。",
        "ip人数": "当日独立 IP 数（含登录+注册 IP）。来源：channel_data_statistics.ip_number。注意：同一 IP 多人登录只算一个，所以 ip人数 > 活跃人数。",
        "充值": "当日所有成功充值订单的总金额（₹）。来源：channel_data_statistics.recharge_amount。等价于 player_recharge_order 表 WHERE order_status=1 的 SUM(pay_amount)。",
        "提现": "当日所有成功提现订单的总金额（₹）。来源：channel_data_statistics.withdraw_amount。等价于 player_withdraw_order 表 WHERE order_status=3 的 SUM(withdraw_amount)，按 finish_time 分日。",
        "充提差": "充值 - 提现。正数=平台净流入，负数=净流出。这是最核心的盈利健康指标。",
        "总投注": "当日所有游戏的总投注金额（₹）。来源：channel_game_statistics.bet_amount，按 statistics_day 汇总所有渠道和游戏类型。注意：不是 player_statistics_day（两表有小差异，以 channel_game_statistics 为准）。",
        "首充人数": "当日首次充值的玩家数（全历史首充，不是当月首充）。来源：channel_data_statistics.first_recharge_number。等价于 player_recharge_order 表 MIN(pay_date) per player。",
        "充值人数": "当日有成功充值的独立玩家数。来源：channel_data_statistics.recharge_number。等价于 player_recharge_order 表 WHERE order_status=1 的 COUNT(DISTINCT player_id)。",
        "人均充值": "充值 ÷ 充值人数。反映付费玩家的平均充值深度。人均越高说明大户贡献越大。",
    }
    
    for name, data in BASE.items():
        ws.cell(r, 1, name)
        if data["total"] is not None:
            c = ws.cell(r, 2, round(data["total"]))
            c.number_format = money
        ws.cell(r, 3, round(data["daily"], 2)).number_format = money2
        ws.cell(r, 4, sources.get(name, ""))
        ws.cell(r, 5, notes.get(name, ""))
        r += 1
    borders(ws, hr, r-1, 1, 5)
    
    # 增长假设
    r += 1
    ws.cell(r, 1, "增长假设").font = hfont
    r += 1
    for m, g in GROWTH.items():
        ws.cell(r, 1, m)
        ws.cell(r, 2, f"{g*100:.0f}%")
        r += 1
    
    # 月度预期总览
    r += 1
    ws.cell(r, 1, "月度预期总览（日均 × 天数 × (1+增长)）").font = hfont
    r += 1
    mheaders = ["月份", "天数", "增长", "预期注册", "预期充值", "预期投注", "预期首充", "预期充值人数"]
    for i, h in enumerate(mheaders, 1):
        c = ws.cell(r, i, h)
        c.font = hfont
        c.fill = blue
    mhr = r
    r += 1
    
    # 预期指标口径说明行
    target_notes = {
        4: "= 注册人数日均 × 天数 × (1+增长率)。口径：channel_data_statistics.register_number",
        5: "= 充值日均 × 天数 × (1+增长率)。口径：channel_data_statistics.recharge_amount（order_status=1 成功订单）",
        6: "= 总投注日均 × 天数 × (1+增长率)。口径：channel_game_statistics.bet_amount（⚠️不用player_statistics_day）",
        7: "= 首充人数日均 × 天数 × (1+增长率)。口径：channel_data_statistics.first_recharge_number（全历史首充）",
        8: "= 充值人数日均 × 天数 × (1+增长率)。口径：channel_data_statistics.recharge_number（当日有成功充值的独立玩家数）",
    }
    
    months_info = [
        ("2026-02", 2, 28),
        ("2026-03", 3, 31),
        ("2026-04", 4, 30),
    ]
    for label, month, days in months_info:
        mname = f"{month}月"
        g = GROWTH[mname]
        ws.cell(r, 1, label)
        ws.cell(r, 2, days)
        ws.cell(r, 3, f"{g*100:.0f}%")
        ws.cell(r, 4, round(BASE["注册人数"]["daily"] * days * (1+g))).number_format = money
        ws.cell(r, 5, round(BASE["充值"]["daily"] * days * (1+g))).number_format = money
        ws.cell(r, 6, round(BASE["总投注"]["daily"] * days * (1+g))).number_format = money
        ws.cell(r, 7, round(BASE["首充人数"]["daily"] * days * (1+g))).number_format = money
        ws.cell(r, 8, round(BASE["充值人数"]["daily"] * days * (1+g))).number_format = money
        r += 1
    borders(ws, mhr, r-1, 1, 8)
    
    # 预期指标口径说明
    r += 1
    ws.cell(r, 1, "预期指标口径说明").font = hfont
    r += 1
    for col_idx, note in target_notes.items():
        header_name = mheaders[col_idx - 1]
        ws.cell(r, 1, header_name).font = Font(bold=True)
        ws.cell(r, 2, note)
        r += 1
    r += 1
    ws.cell(r, 1, "公式").font = Font(bold=True)
    ws.cell(r, 2, "预期值 = 基准日均 x 当月天数 x (1 + 月增长率)。基准日均来自 12/1~1/30 共 61 天 DB 实际数据。")
    
    for col, w in [("A", 18), ("B", 18), ("C", 14), ("D", 22), ("E", 60), ("F", 18), ("G", 16), ("H", 16)]:
        ws.column_dimensions[col].width = w
    
    # === Monthly sheets with weekly checkpoints ===
    metrics_order = ["注册人数", "活跃人数", "ip人数", "充值", "提现", "充提差", "总投注", "首充人数", "充值人数", "人均充值"]
    
    for label, month, days in months_info:
        mname = f"{month}月"
        g = GROWTH[mname]
        ws2 = wb.create_sheet(title=mname)
        
        ws2.merge_cells("A1:L1")
        ws2["A1"] = f"2026 {mname} 周预期进度（周日结算）"
        ws2["A1"].font = tfont
        ws2["A2"] = f"增长假设: {g*100:.0f}% | 基准: 12月+1月日均 × (1+{g*100:.0f}%)"
        ws2["A2"].font = sfont
        
        # Get weekly checkpoints
        sundays = get_weekly_sundays(2026, month)
        
        # Header row
        cols = ["周结算日", "累计天数", "时间进度"] + metrics_order
        for i, h in enumerate(cols, 1):
            c = ws2.cell(4, i, h)
            c.font = hfont
            c.fill = blue
        
        first_day = date(2026, month, 1)
        
        for j, sunday in enumerate(sundays, 5):
            cum_days = (sunday - first_day).days + 1
            time_pct = cum_days / days
            
            ws2.cell(j, 1, sunday.strftime("%Y-%m-%d"))
            ws2.cell(j, 2, cum_days)
            ws2.cell(j, 3, time_pct).number_format = pct
            
            for k, metric in enumerate(metrics_order, 4):
                daily = BASE[metric]["daily"]
                if metric == "人均充值":
                    # 人均充值 doesn't scale by days
                    val = round(daily * (1+g), 2)
                elif metric == "充提差":
                    val = round(BASE["充值"]["daily"] * cum_days * (1+g) - BASE["提现"]["daily"] * cum_days * (1+g))
                else:
                    val = round(daily * cum_days * (1+g))
                
                c = ws2.cell(j, k, val)
                c.number_format = money if metric != "人均充值" else money2
        
        borders(ws2, 4, 4+len(sundays), 1, len(cols))
        
        # Month total row
        tr = 5 + len(sundays) + 1
        ws2.cell(tr, 1, "月度总目标").font = Font(bold=True, size=12)
        for k, metric in enumerate(metrics_order, 4):
            daily = BASE[metric]["daily"]
            if metric == "人均充值":
                val = round(daily * (1+g), 2)
            elif metric == "充提差":
                val = round(BASE["充值"]["daily"] * days * (1+g) - BASE["提现"]["daily"] * days * (1+g))
            else:
                val = round(daily * days * (1+g))
            c = ws2.cell(tr, k, val)
            c.number_format = money if metric != "人均充值" else money2
            c.font = Font(bold=True)
        
        # Actual row (placeholder)
        tr2 = tr + 1
        ws2.cell(tr2, 1, "实际累计").font = Font(bold=True, color="FF0000")
        
        # Diff row
        tr3 = tr2 + 1
        ws2.cell(tr3, 1, "差异").font = Font(bold=True, color="FF0000")
        
        # Column widths
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 10
        for i, m in enumerate(metrics_order, 4):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws2.column_dimensions[col_letter].width = 16
    
    # === 口径说明 Sheet ===
    ws_notes = wb.create_sheet(title="口径说明")
    ws_notes.merge_cells("A1:C1")
    ws_notes["A1"] = "BG666 报表口径说明（2026-01-31 对齐确认）"
    ws_notes["A1"].font = tfont
    
    ws_notes["A2"] = "所有指标已与后台 Excel 报表逐日对齐验证（12/27-1/3，8天全部 ✅）"
    ws_notes["A2"].font = sfont
    
    note_headers = ["指标", "DB 来源表", "DB 字段", "SQL 口径", "备注"]
    for i, h in enumerate(note_headers, 1):
        c = ws_notes.cell(4, i, h)
        c.font = hfont
        c.fill = blue
    
    note_rows = [
        ("注册人数", "channel_data_statistics", "register_number",
         "SUM(register_number) GROUP BY statistics_day",
         "当日新注册玩家数，汇总所有渠道。与 sys_player COUNT(create_time) 一致。"),
        ("活跃人数", "channel_data_statistics", "active_number",
         "SUM(active_number) GROUP BY statistics_day",
         "当日至少登录一次的独立玩家数。与 player_logininfor COUNT(DISTINCT player_id) 一致。"),
        ("ip人数", "channel_data_statistics", "ip_number",
         "SUM(ip_number) GROUP BY statistics_day",
         "当日独立 IP 数（含登录+注册）。注意 ip人数 > 活跃人数，因为同一 IP 可能有未注册访问。不能从 player_logininfor DISTINCT ipaddr 得到（那个偏小约3600）。"),
        ("充值", "channel_data_statistics", "recharge_amount",
         "SUM(recharge_amount) GROUP BY statistics_day",
         "当日成功充值订单总金额（₹）。与 player_recharge_order WHERE order_status=1 的 SUM(pay_amount) BY pay_date 一致。"),
        ("提现", "channel_data_statistics", "withdraw_amount",
         "SUM(withdraw_amount) GROUP BY statistics_day",
         "当日成功提现订单总金额（₹）。与 player_withdraw_order WHERE order_status=3 的 SUM(withdraw_amount) BY finish_time 一致。注意：提现成功状态是 3（不是 1）。"),
        ("充提差", "计算字段", "recharge - withdraw",
         "SUM(recharge_amount) - SUM(withdraw_amount)",
         "充值减提现。正=净流入，负=净流出。12/31 曾出现负值（-259,691）。核心健康指标。"),
        ("总投注", "channel_game_statistics", "bet_amount",
         "SUM(bet_amount) GROUP BY statistics_day",
         "当日所有游戏总投注（₹），汇总所有渠道和游戏类型。⚠️ 不要用 player_statistics_day.bet_amount（有小差异，如12/27差10K）。以本表为准。"),
        ("首充人数", "channel_data_statistics", "first_recharge_number",
         "SUM(first_recharge_number) GROUP BY statistics_day",
         "当日首次充值的玩家数（全历史首充）。与 player_recharge_order MIN(pay_date) per player 一致。⚠️ 不要用 first_deposit_record 表（二级数据，数字不准）。"),
        ("充值人数", "channel_data_statistics", "recharge_number",
         "SUM(recharge_number) GROUP BY statistics_day",
         "当日有成功充值的独立玩家数。与 player_recharge_order WHERE order_status=1 的 COUNT(DISTINCT player_id) BY pay_date 一致。"),
        ("人均充值", "计算字段", "充值 / 充值人数",
         "SUM(recharge_amount) / SUM(recharge_number)",
         "付费玩家平均充值深度（₹）。人均越高=大户贡献越大。"),
    ]
    
    for j, row in enumerate(note_rows, 5):
        for i, val in enumerate(row, 1):
            ws_notes.cell(j, i, val)
    borders(ws_notes, 4, 4+len(note_rows), 1, 5)
    
    # 补充说明
    nr = 5 + len(note_rows) + 2
    ws_notes.cell(nr, 1, "重要补充").font = hfont
    nr += 1
    supplements = [
        "1. 数据来源：channel_data_statistics（8个指标）+ channel_game_statistics（总投注）= 两张表覆盖全部 9 个指标。",
        "2. 基准期：2025-12-01 ~ 2026-01-30，共 61 天。",
        "3. 预期计算：日均 × 天数 × (1 + 增长率)。增长率统一 5%。",
        "4. 周结算日：每周日 + 月末（如果月末不是周日则额外加入）。",
        "5. 充提差中的提现口径：order_status=3 表示成功（不是 1），时间用 finish_time（完成时间，不是 apply_time）。",
        "6. 总投注必须用 channel_game_statistics，不能用 player_statistics_day（后者12/27差10K，12/29差350，1/3差160）。",
        "7. ip人数必须用 channel_data_statistics.ip_number，不能从 player_logininfor DISTINCT ipaddr 推算（偏小约3600-4400/天）。",
        "8. 首充必须从 player_recharge_order MIN(pay_date) 推导，不能用 first_deposit_record 表（二级数据）。",
    ]
    for s in supplements:
        ws_notes.cell(nr, 1, s)
        nr += 1
    
    ws_notes.column_dimensions["A"].width = 14
    ws_notes.column_dimensions["B"].width = 28
    ws_notes.column_dimensions["C"].width = 26
    ws_notes.column_dimensions["D"].width = 50
    ws_notes.column_dimensions["E"].width = 80
    
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"Done: {OUTPUT}")

if __name__ == "__main__":
    main()
