#!/usr/bin/env python3
"""Generate BG666 2-4月 预期目标 Excel from PDF data."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from pathlib import Path

OUTPUT = Path(__file__).parent.parent.parent.parent / "output" / "bg666_2026_feb_target_plan_sc.xlsx"

def main():
    wb = openpyxl.Workbook()
    
    hfont = Font(bold=True, size=11)
    tfont = Font(bold=True, size=14)
    sfont = Font(bold=True, size=11, color="666666")
    blue = PatternFill(start_color="D6EAF8", fill_type="solid")
    yellow = PatternFill(start_color="FEF9E7", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    money = "#,##0"
    pct = "0.0%"
    
    def borders(ws, r1, r2, c1, c2):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).border = border

    # === Sheet 1: Overview ===
    ws = wb.active
    ws.title = "总览"
    
    ws.merge_cells("A1:F1")
    ws["A1"].font = tfont
    ws["A1"] = "BG666 2026 2-4月 预期进度表"
    ws["A2"] = "基于 2025-12 与 2026-01 的注册、首充、流失模型推算（保守版）"
    ws["A2"].font = sfont
    ws["A3"] = "口径：自然日 | 周结算：周日 | 核心KPI：存款/进度差"
    ws["A4"] = "基准期：2025-12-01 ~ 2026-01-30（61天）"
    
    # Base metrics
    r = 6
    ws.cell(r, 1, "基准指标").font = hfont
    r += 1
    base_metrics = [
        ("日均注册", 1016),
        ("日均存款", 3998636),
        ("日均提现", 3471051),
        ("首充率", 0.718),
        ("首充均额", 636),
        ("付费用户月流失率", 0.467),
    ]
    for name, val in base_metrics:
        ws.cell(r, 1, name).font = Font(bold=True)
        c = ws.cell(r, 2, val)
        if isinstance(val, float) and val < 1:
            c.number_format = pct
        else:
            c.number_format = money
        r += 1
    
    # Growth assumptions
    r += 1
    ws.cell(r, 1, "增长假设").font = hfont
    r += 1
    for m, p in [("2月", "3%"), ("3月", "4%"), ("4月", "5%")]:
        ws.cell(r, 1, m)
        ws.cell(r, 2, p)
        r += 1
    
    # Monthly summary
    r += 1
    ws.cell(r, 1, "月度预期总览").font = hfont
    r += 1
    headers = ["月份", "增长假设", "预期注册量", "预期首充人數", "预期存款", "预期投注额"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h)
        c.font = hfont
        c.fill = blue
    hr = r
    r += 1
    
    monthly = [
        ("2026-02", "3%", 29308, 21056, 112878928, 938142562),
        ("2026-03", "4%", 32763, 23538, 106823125, 887812468),
        ("2026-04", "5%", 32011, 22998, 106982407, 889136268),
    ]
    for row_data in monthly:
        for i, val in enumerate(row_data, 1):
            c = ws.cell(r, i, val)
            if i >= 3:
                c.number_format = money
        r += 1
    borders(ws, hr, r-1, 1, 6)
    
    # Formula
    r += 1
    ws.cell(r, 1, "预期存款 = 预期注册 × 首充率 × 首充均额 + (上月付费用户 × (1-流失率) × 复充均额)").font = sfont
    
    # Events
    r += 2
    ws.cell(r, 1, "事件表（影响投注额的关键节点）").font = hfont
    r += 1
    evt_h = ["事件", "日期", "影响假设", "备注"]
    for i, h in enumerate(evt_h, 1):
        c = ws.cell(r, i, h)
        c.font = hfont
        c.fill = yellow
    er = r
    r += 1
    events = [
        ("ICC Men's T20 World Cup 2026", "2026-02-07 ~ 2026-03-08", "投注额显著上升（主影响）", "uplift 可调"),
        ("FIFA World Cup 26", "2026-06-11 ~ 2026-07-19", "投注额显著上升（淘汰赛最強）", "不在2-4月区间，仅保留参考"),
    ]
    for evt in events:
        for i, val in enumerate(evt, 1):
            ws.cell(r, i, val)
        r += 1
    borders(ws, er, r-1, 1, 4)
    
    for col, w in [("A", 30), ("B", 25), ("C", 20), ("D", 18), ("E", 18), ("F", 18)]:
        ws.column_dimensions[col].width = w
    
    # === Monthly sheets ===
    detail = {
        "2月": [
            ("2026-02-01", 3.6, 4031390),
            ("2026-02-08", 28.6, 32251122),
            ("2026-02-15", 53.6, 60470855),
            ("2026-02-22", 78.6, 88690587),
            ("2026-02-28", 100.0, 112878928),
        ],
        "3月": [
            ("2026-03-01", 3.2, 3445907),
            ("2026-03-08", 25.8, 27567258),
            ("2026-03-15", 48.4, 51688609),
            ("2026-03-22", 71.0, 75809960),
            ("2026-03-29", 93.5, 99931310),
            ("2026-03-31", 100.0, 106823125),
        ],
        "4月": [
            ("2026-04-05", 16.7, 17830401),
            ("2026-04-12", 40.0, 42792963),
            ("2026-04-19", 63.3, 67755524),
            ("2026-04-26", 86.7, 92718086),
            ("2026-04-30", 100.0, 106982407),
        ],
    }
    totals = {"2月": 112878928, "3月": 106823125, "4月": 106982407}
    
    for mname, checkpoints in detail.items():
        ws2 = wb.create_sheet(title=mname)
        ws2.merge_cells("A1:G1")
        ws2["A1"] = f"2026 {mname} 每周进度预期（周日结算）"
        ws2["A1"].font = tfont
        
        cols = ["周结算日", "時間进度", "存款目标", "预期存款", "进度差(pp)", "实际存款", "实际进度差"]
        for i, h in enumerate(cols, 1):
            c = ws2.cell(3, i, h)
            c.font = hfont
            c.fill = blue
        
        for j, (dt, p, target) in enumerate(checkpoints, 4):
            ws2.cell(j, 1, dt)
            c = ws2.cell(j, 2, p/100)
            c.number_format = pct
            c = ws2.cell(j, 3, target)
            c.number_format = money
            c = ws2.cell(j, 4, target)
            c.number_format = money
            ws2.cell(j, 5, 0).number_format = "+0.0;-0.0"
            ws2.cell(j, 6, "").number_format = money  # placeholder
            ws2.cell(j, 7, "")  # placeholder
        
        borders(ws2, 3, 3+len(checkpoints), 1, 7)
        
        r2 = 4 + len(checkpoints) + 1
        ws2.cell(r2, 1, "月度总目标").font = hfont
        c = ws2.cell(r2, 3, totals[mname])
        c.number_format = money
        c.font = Font(bold=True, size=12)
        
        for col, w in [("A", 16), ("B", 14), ("C", 18), ("D", 18), ("E", 14), ("F", 18), ("G", 14)]:
            ws2.column_dimensions[col].width = w
    
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"Done: {OUTPUT}")

if __name__ == "__main__":
    main()
