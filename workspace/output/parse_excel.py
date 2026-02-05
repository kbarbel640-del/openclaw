import openpyxl
import json

wb = openpyxl.load_workbook('/Users/sulaxd/clawd/output/feb_target.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== {name} ({ws.max_row}r x {ws.max_column}c) ===')
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=False):
        vals = [(c.coordinate, c.value) for c in row if c.value is not None]
        if vals:
            print(vals)
