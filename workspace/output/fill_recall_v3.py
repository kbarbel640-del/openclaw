#!/usr/bin/env python3
"""Fill Red's recall Excel using player_id from her new files."""
import subprocess, json, sys
from datetime import date, timedelta

QUERY = "/Users/sulaxd/clawd/skills/bg666-db/scripts/query.py"
DL = "/Users/sulaxd/clawd/skills/telegram-userbot/downloads/"
INPUT = DL + "活动分析需求2.1xlsx.xlsx"
OUTPUT = "/Users/sulaxd/clawd/output/recall_activity_filled_v3.xlsx"

# Red's new files with login_name -> player_id mapping
MAPPING_FILES = [
    ("标签1（＜500）.xlsx", "发送会员账号"),
    ("标签1（≥500）.xlsx", "发送会员账号"),
    ("标签2（第二批M）.xlsx", "发送会员账号"),
    ("标签2（第一批C）.xlsx", "发送会员账号"),
]

def query(sql, timeout=60):
    r = subprocess.run(["python3", QUERY, "--json", sql],
                       capture_output=True, text=True, timeout=timeout)
    if r.returncode != 0:
        print(f"  Query error: {r.stderr[:200]}", file=sys.stderr)
        return []
    return json.loads(r.stdout)

def serial_to_date(serial):
    base = date(1899, 12, 30)
    return (base + timedelta(days=int(serial))).strftime("%Y-%m-%d")

def main():
    import openpyxl

    # Step 1: Build login_name -> player_id mapping from Red's new files
    login_to_pid = {}
    for fname, sheet in MAPPING_FILES:
        wb2 = openpyxl.load_workbook(DL + fname, read_only=True)
        ws2 = wb2[sheet]
        for row in ws2.iter_rows(min_row=2, max_col=2, values_only=True):
            login, pid = row
            if login is not None and pid is not None:
                try:
                    login_to_pid[str(login).strip()] = int(pid)
                except (ValueError, TypeError):
                    pass  # skip #N/A or invalid entries
        wb2.close()
    print(f"Mapping loaded: {len(login_to_pid)} entries")

    # Step 2: Open main Excel, map Col A to player_id
    wb = openpyxl.load_workbook(INPUT)
    
    all_pids = set()
    sheet_mapping = {}  # (sheet, row) -> player_id
    unmapped = 0
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(4, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            key = str(v).strip()
            pid = login_to_pid.get(key)
            if pid:
                sheet_mapping[(sname, r)] = pid
                all_pids.add(pid)
            else:
                # Maybe it's already a player_id?
                try:
                    pid = int(v)
                    if pid < 200000000:  # reasonable player_id range
                        sheet_mapping[(sname, r)] = pid
                        all_pids.add(pid)
                    else:
                        unmapped += 1
                except:
                    unmapped += 1
    
    print(f"Mapped players: {len(all_pids)}, unmapped: {unmapped}")

    # Step 3: Parse date columns
    ws1 = wb[wb.sheetnames[0]]
    date_cols = {}
    for c in range(8, 24):
        v = ws1.cell(3, c).value
        if v:
            try:
                ds = serial_to_date(int(v))
                date_cols[c] = ds
            except:
                pass
    print(f"Date columns: {list(date_cols.values())}")

    # Step 4: Batch query player info
    pid_list = sorted(all_pids)
    player_info = {}
    batch = 500
    
    print("Loading player info...")
    for i in range(0, len(pid_list), batch):
        b = pid_list[i:i+batch]
        pids_str = ",".join(str(p) for p in b)
        rows = query(f"SELECT player_id, nick_name, vip_level, first_login_day FROM sys_player WHERE player_id IN ({pids_str})")
        for row in rows:
            player_info[int(row['player_id'])] = row
        print(f"  Batch {i//batch+1}: {len(rows)} rows")
    print(f"Player info: {len(player_info)}")

    # Step 5: Cumulative stats
    player_stats = {}
    print("Loading cumulative stats...")
    for i in range(0, len(pid_list), batch):
        b = pid_list[i:i+batch]
        pids_str = ",".join(str(p) for p in b)
        rows = query(f"""
            SELECT player_id,
                   SUM(recharge_amount) as total_recharge,
                   SUM(bet_amount) as total_bet,
                   SUM(win_amount - bet_amount) as total_winloss
            FROM player_statistics_day
            WHERE player_id IN ({pids_str})
            GROUP BY player_id
        """, timeout=120)
        for row in rows:
            player_stats[int(row['player_id'])] = row
        print(f"  Batch {i//batch+1}: {len(rows)} rows")
    print(f"Player stats: {len(player_stats)}")

    # Step 6: Daily deposits
    daily_deps = {}
    all_dates = sorted(date_cols.values())
    date_min, date_max = all_dates[0], all_dates[-1]
    print(f"Loading daily deposits {date_min} to {date_max}...")
    for i in range(0, len(pid_list), batch):
        b = pid_list[i:i+batch]
        pids_str = ",".join(str(p) for p in b)
        rows = query(f"""
            SELECT player_id, DATE(create_time) as dt, SUM(pay_amount) as dep
            FROM player_recharge_order
            WHERE player_id IN ({pids_str})
              AND order_status = 1
              AND DATE(create_time) BETWEEN '{date_min}' AND '{date_max}'
            GROUP BY player_id, DATE(create_time)
        """, timeout=120)
        for row in rows:
            daily_deps[(int(row['player_id']), str(row['dt']))] = float(row['dep'])
        print(f"  Batch {i//batch+1}: {len(rows)} rows")
    print(f"Daily deposit records: {len(daily_deps)}")

    # Step 7: Fill Excel
    for sname in wb.sheetnames:
        ws = wb[sname]
        filled = 0
        for r in range(4, ws.max_row + 1):
            pid = sheet_mapping.get((sname, r))
            if not pid:
                continue
            
            info = player_info.get(pid, {})
            stats = player_stats.get(pid, {})
            
            # Col A: nick_name (replace login_name)
            ws.cell(r, 1).value = info.get('nick_name', ws.cell(r, 1).value)
            # Col B: player_id
            ws.cell(r, 2).value = pid
            # Col C: VIP
            ws.cell(r, 3).value = info.get('vip_level', 0)
            # Col D: register time
            ws.cell(r, 4).value = str(info.get('first_login_day', '') or '')
            # Col E: cumulative recharge
            ws.cell(r, 5).value = float(stats.get('total_recharge', 0) or 0)
            # Col F: cumulative bet
            ws.cell(r, 6).value = float(stats.get('total_bet', 0) or 0)
            # Col G: total win/loss
            ws.cell(r, 7).value = float(stats.get('total_winloss', 0) or 0)
            
            # Daily deposits
            for c, ds in date_cols.items():
                ws.cell(r, c).value = daily_deps.get((pid, ds), 0)
            
            filled += 1
        print(f"Sheet '{sname}': filled {filled} rows")

    wb.save(OUTPUT)
    print(f"\n✅ Saved to {OUTPUT}")

if __name__ == "__main__":
    main()
