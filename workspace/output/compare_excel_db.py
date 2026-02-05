import openpyxl
wb = openpyxl.load_workbook('/Users/sulaxd/.moltbot/media/inbound/file_555---2033aa34-eb39-42de-9f34-58dad20ffe0e.xlsx', data_only=True)

# 12月: 充值=col7, 首充人数=col15, 活跃=col5
ws = wb['12月 (2)']
print('=== 12月 Excel (sample) ===')
print('日期 | 充值 | 首充 | 活躍')
for r in range(2, 7):
    dt = str(ws.cell(r,1).value)[:10]
    dep = ws.cell(r,7).value
    fd = ws.cell(r,15).value
    active = ws.cell(r,5).value
    print(f'{dt} | {dep} | {fd} | {active}')

# 1月: 充值=col7, 首充=col15, 充值人数=col16, 活跃=col5
ws2 = wb['1月']
print()
print('=== 1月 Excel (sample) ===')
print('日期 | 充值 | 首充 | 充值人數 | 活躍')
for r in range(2, 7):
    dt = str(ws2.cell(r,1).value)[:10]
    dep = ws2.cell(r,7).value
    fd = ws2.cell(r,15).value
    dp = ws2.cell(r,16).value
    active = ws2.cell(r,5).value
    print(f'{dt} | {dep} | {fd} | {dp} | {active}')

# 12月 totals
print()
print('=== 12月 Excel 月合計 ===')
total_dep = sum(ws.cell(r,7).value or 0 for r in range(2, ws.max_row+1))
total_fd = sum(ws.cell(r,15).value or 0 for r in range(2, ws.max_row+1))
print(f'充值總額: {total_dep}')
print(f'首充人數總計: {total_fd}')
print(f'天數: {ws.max_row-1}')

# 1月 totals
print()
print('=== 1月 Excel 月合計 ===')
total_dep2 = sum(ws2.cell(r,7).value or 0 for r in range(2, ws2.max_row+1))
total_fd2 = sum(ws2.cell(r,15).value or 0 for r in range(2, ws2.max_row+1))
total_dp2 = sum(ws2.cell(r,16).value or 0 for r in range(2, ws2.max_row+1))
print(f'充值總額: {total_dep2}')
print(f'首充人數總計: {total_fd2}')
print(f'充值人數總計(日累加): {total_dp2}')
print(f'天數: {ws2.max_row-1}')
