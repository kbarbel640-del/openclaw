#!/usr/bin/env python3
"""Fill Red's recall activity Excel with DB data."""
import subprocess, json, sys, os
from datetime import datetime, timedelta

QUERY_SCRIPT = "/Users/sulaxd/clawd/skills/bg666-db/scripts/query.py"
INPUT_FILE = "/Users/sulaxd/clawd/skills/telegram-userbot/downloads/活动分析需求2.1xlsx.xlsx"
OUTPUT_FILE = "/Users/sulaxd/clawd/output/recall_activity_filled_v2.xlsx"

def query(sql, timeout=30):
    """Run DB query via query.py --json"""
    r = subprocess.run(
        ["python3", QUERY_SCRIPT, "--json", sql],
        capture_output=True, text=True, timeout=timeout
    )
    if r.returncode != 0:
        print(f"Query error: {r.stderr[:200]}", file=sys.stderr)
        return []
    return json.loads(r.stdout)

def serial_to_date(serial):
    """Excel serial number to date string YYYY-MM-DD"""
    # Excel epoch: 1899-12-30
    from datetime import date, timedelta
    base = date(1899, 12, 30)
    d = base + timedelta(days=int(serial))
    return d.strftime("%Y-%m-%d")

def main():
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_FILE)
    
    # Collect all player_ids from all sheets
    all_pids = set()
    sheet_pids = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        pids = []
        for r in range(4, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None:
                try:
                    pid = int(v)
                    if pid > 5000000000 and pid < 7000000000:  # player_id range
                        pids.append(pid)
                        all_pids.add(pid)
                except (ValueError, TypeError):
                    pass  # skip login_name entries (phone/email)
        sheet_pids[sname] = pids
    
    print(f"Total unique players: {len(all_pids)}")
    
    # Parse date columns from Row 3 headers
    ws1 = wb[wb.sheetnames[0]]
    date_cols = {}  # col_idx -> date_str
    for c in range(8, ws1.max_column + 1):
        v = ws1.cell(3, c).value
        if v is not None:
            try:
                ds = serial_to_date(int(v))
                date_cols[c] = ds
                print(f"  Col {c} -> {ds}")
            except:
                pass
    
    # Build date list for daily deposit query
    all_dates = sorted(set(date_cols.values()))
    print(f"Date range: {all_dates[0]} to {all_dates[-1]}")
    
    # Step 1: Get player info (batch by 500)
    pid_list = sorted(all_pids)
    player_info = {}  # pid -> {nick_name, vip_level, first_login_day, ...}
    
    batch_size = 500
    for i in range(0, len(pid_list), batch_size):
        batch = pid_list[i:i+batch_size]
        pids_str = ",".join(str(p) for p in batch)
        rows = query(f"""
            SELECT player_id, nick_name, vip_level, first_login_day
            FROM sys_player 
            WHERE player_id IN ({pids_str})
        """, timeout=60)
        for row in rows:
            player_info[int(row['player_id'])] = row
        print(f"  Player info batch {i//batch_size+1}: {len(rows)} rows")
    
    print(f"Player info loaded: {len(player_info)}")
    
    # Step 2: Get cumulative stats (recharge, bet, win/loss)
    # Use player_statistics_day to sum all-time
    player_stats = {}
    for i in range(0, len(pid_list), batch_size):
        batch = pid_list[i:i+batch_size]
        pids_str = ",".join(str(p) for p in batch)
        rows = query(f"""
            SELECT player_id, 
                   SUM(recharge_amount) as total_recharge,
                   SUM(bet_amount) as total_bet,
                   SUM(win_amount - bet_amount) as total_winloss
            FROM player_statistics_day
            WHERE player_id IN ({pids_str})
            GROUP BY player_id
        """, timeout=120)
        for row in rows:
            player_stats[int(row['player_id'])] = row
        print(f"  Stats batch {i//batch_size+1}: {len(rows)} rows")
    
    print(f"Player stats loaded: {len(player_stats)}")
    
    # Step 3: Get daily deposits for date range
    daily_deposits = {}  # (pid, date) -> amount
    date_min = all_dates[0]
    date_max = all_dates[-1]
    
    for i in range(0, len(pid_list), batch_size):
        batch = pid_list[i:i+batch_size]
        pids_str = ",".join(str(p) for p in batch)
        rows = query(f"""
            SELECT player_id, DATE(create_time) as dt, SUM(amount) as daily_dep
            FROM player_recharge_order
            WHERE player_id IN ({pids_str})
              AND pay_status = 1
              AND DATE(create_time) BETWEEN '{date_min}' AND '{date_max}'
            GROUP BY player_id, DATE(create_time)
        """, timeout=120)
        for row in rows:
            key = (int(row['player_id']), str(row['dt']))
            daily_deposits[key] = float(row['daily_dep'])
        print(f"  Daily deposits batch {i//batch_size+1}: {len(rows)} rows")
    
    print(f"Daily deposit records: {len(daily_deposits)}")
    
    # Step 4: Fill the Excel (only sheets with player_ids)
    for sname in wb.sheetnames:
        ws = wb[sname]
        filled = 0
        for r in range(4, ws.max_row + 1):
            pid_val = ws.cell(r, 1).value
            if pid_val is None:
                continue
            try:
                pid = int(pid_val)
                if pid < 5000000000 or pid > 7000000000:
                    continue  # skip phone/email rows
            except (ValueError, TypeError):
                continue
            
            info = player_info.get(pid, {})
            stats = player_stats.get(pid, {})
            
            # Col A: 会员账号 -> nick_name
            ws.cell(r, 1).value = info.get('nick_name', pid_val)
            # Col B: 会员ID -> player_id
            ws.cell(r, 2).value = pid
            # Col C: VIP等级
            ws.cell(r, 3).value = info.get('vip_level', 0)
            # Col D: 账号注册时间
            ws.cell(r, 4).value = info.get('first_login_day', '')
            # Col E: 历史累计存款
            ws.cell(r, 5).value = float(stats.get('total_recharge', 0) or 0)
            # Col F: 历史累计投注
            ws.cell(r, 6).value = float(stats.get('total_bet', 0) or 0)
            # Col G: 历史总输赢
            ws.cell(r, 7).value = float(stats.get('total_winloss', 0) or 0)
            
            # Daily deposits
            for c, ds in date_cols.items():
                amt = daily_deposits.get((pid, ds), 0)
                ws.cell(r, c).value = amt if amt else 0
            
            filled += 1
        
        print(f"Sheet '{sname}': filled {filled} rows")
    
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
