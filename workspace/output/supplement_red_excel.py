#!/usr/bin/env python3
"""Supplement Red's Excel with 1/31 withdraw amount and last login time."""
import subprocess, json, sys
from datetime import date, timedelta

QUERY = "/Users/sulaxd/clawd/skills/bg666-db/scripts/query.py"
INPUT = "/Users/sulaxd/clawd/output/recall_activity_filled_v3.xlsx"
OUTPUT = "/Users/sulaxd/clawd/output/recall_activity_filled_v4.xlsx"

def query(sql, timeout=60):
    r = subprocess.run(["python3", QUERY, "--json", sql],
                       capture_output=True, text=True, timeout=timeout)
    if r.returncode != 0:
        print(f"Query error: {r.stderr[:200]}", file=sys.stderr)
        return []
    return json.loads(r.stdout)

def main():
    import openpyxl
    
    wb = openpyxl.load_workbook(INPUT)
    
    # Collect all player_ids from column B
    all_pids = set()
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(4, ws.max_row + 1):
            pid = ws.cell(r, 2).value  # Col B = player_id
            if pid:
                all_pids.add(int(pid))
    
    print(f"Total players: {len(all_pids)}")
    
    # Get last_login_day from sys_player
    pid_list = sorted(all_pids)
    last_login = {}
    batch = 500
    
    print("Loading last login times...")
    for i in range(0, len(pid_list), batch):
        b = pid_list[i:i+batch]
        pids_str = ",".join(str(p) for p in b)
        rows = query(f"SELECT player_id, last_login_day FROM sys_player WHERE player_id IN ({pids_str})")
        for row in rows:
            last_login[int(row['player_id'])] = row.get('last_login_day', '')
        print(f"  Batch {i//batch+1}: {len(rows)} rows")
    
    print(f"Last login loaded: {len(last_login)}")
    
    # Get 1/31 withdraw amounts
    print("Loading 1/31 withdraw amounts...")
    withdraw_131 = {}
    for i in range(0, len(pid_list), batch):
        b = pid_list[i:i+batch]
        pids_str = ",".join(str(p) for p in b)
        rows = query(f"""
            SELECT player_id, SUM(withdraw_amount) as withdraw_amt
            FROM player_withdraw_order
            WHERE player_id IN ({pids_str})
              AND order_status = 3
              AND DATE(create_time) = '2026-01-31'
            GROUP BY player_id
        """, timeout=120)
        for row in rows:
            withdraw_131[int(row['player_id'])] = float(row['withdraw_amt'])
        print(f"  Batch {i//batch+1}: {len(rows)} rows")
    
    print(f"1/31 withdrawals loaded: {len(withdraw_131)}")
    
    # Find where to add new columns
    # Current structure: A=nick_name, B=player_id, C=VIP, D=register_time, E=cum_recharge, F=cum_bet, G=cum_winloss, H+=daily deposits
    # We need to add: last_login_time after D, and 1/31 withdraw somewhere
    
    # For now, let's add to the end or find the right spot
    # Looking at the Excel, col 15 is 1/31 deposit. Let's add withdraw next to it.
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        filled = 0
        
        # Add headers in row 3
        # Find max column
        max_col = ws.max_column
        
        # Add "最後上線時間" header and "1/31提款" header
        # Let's put last_login after registration (col D), but that would shift everything
        # Easier: add at the end
        ws.cell(3, max_col + 1).value = "最後上線時間"
        ws.cell(3, max_col + 2).value = "1/31提款金額"
        
        for r in range(4, ws.max_row + 1):
            pid = ws.cell(r, 2).value
            if not pid:
                continue
            pid = int(pid)
            
            # Add last login time
            ws.cell(r, max_col + 1).value = str(last_login.get(pid, '') or '')
            
            # Add 1/31 withdraw
            ws.cell(r, max_col + 2).value = withdraw_131.get(pid, 0)
            
            filled += 1
        
        print(f"Sheet '{sname}': supplemented {filled} rows")
    
    wb.save(OUTPUT)
    print(f"\n✅ Saved to {OUTPUT}")

if __name__ == "__main__":
    main()
