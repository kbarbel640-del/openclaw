#!/usr/bin/env python3
"""
Fill Red's recall activity Excel template with data from DB.
4 sheets, each with player lists needing:
- VIP等级, 账号注册时间, 历史累计存款/投注/输赢
- 1/24-1/30 每日存款
- 当日存款 (1/31)
- 2/2-2/8 每日存款 (only 2/1-2/2 available now)
"""
import pymysql
import openpyxl
from datetime import date, datetime, timedelta
import sys

DB = dict(host='bg666-market-readonly.czsks2mguhd5.ap-south-1.rds.amazonaws.com',
          port=3306, user='market', password='hBVoVVm&)aZtW0t6',
          db='ry-cloud', charset='utf8mb4',
          cursorclass=pymysql.cursors.DictCursor, connect_timeout=30, read_timeout=600)

conn = pymysql.connect(**DB)
cur = conn.cursor()

def q(sql, args=None):
    cur.execute(sql, args)
    return cur.fetchall()

# Date serial mapping (Excel date serials in header row)
# 46046 = 2026-01-24, 46047 = 2026-01-25, ...
# Let's verify: Excel serial 1 = 1900-01-01, so 46046 = 2026-01-24
date_cols = {
    7: date(2026,1,24),   # col H (0-indexed: 7)
    8: date(2026,1,25),
    9: date(2026,1,26),
    10: date(2026,1,27),
    11: date(2026,1,28),
    12: date(2026,1,29),
    13: date(2026,1,30),
    14: date(2026,1,31),  # 当日存款
    15: date(2026,2,1),   # Wait - need to check. Let me use serial mapping
    16: date(2026,2,2),
    17: date(2026,2,3),
    18: date(2026,2,4),
    19: date(2026,2,5),
    20: date(2026,2,6),
    21: date(2026,2,7),
    22: date(2026,2,8),
}

# Actually, the date serials 46054-46061 correspond to:
# 46054 = 46046 + 8 = Jan 24 + 8 = Feb 1
# So: col 15 = Feb 1, col 16 = Feb 2, etc.
# But header says "2月2日至2月8日" and "当日存款" is a single column
# Let me re-check: header row 2 says:
# cols 7-13: "1月24日至1月30日每日存款" (7 days)
# col 14: "当日存款" (46053 = Jan 31)  
# cols 15-22: "2月2日至2月8日" but serials are 46054-46061 = Feb 1 to Feb 8

# Actually 46054 = 46046+8 = Jan 24+8 = Feb 1
# So the mapping is correct. Let me adjust:
date_cols_corrected = {}
base_serial = 46046
base_date = date(2026, 1, 24)
for col_idx in range(7, 23):  # H to W (0-indexed 7-22)
    if col_idx <= 13:  # 1/24 - 1/30
        d = base_date + timedelta(days=col_idx - 7)
    elif col_idx == 14:  # 当日存款 = 1/31
        d = date(2026, 1, 31)
    else:  # 2/1 - 2/8
        d = date(2026, 2, 1) + timedelta(days=col_idx - 15)
    date_cols_corrected[col_idx] = d

print("Date column mapping:")
for c, d in date_cols_corrected.items():
    print(f"  Col {c} ({chr(65+c)}): {d}")

# Load workbook
src = '/Users/sulaxd/clawd/skills/telegram-userbot/downloads/活动分析需求2.1xlsx.xlsx'
wb = openpyxl.load_workbook(src)

for sheet_idx, ws in enumerate(wb.worksheets):
    print(f"\n{'='*60}")
    print(f"Processing Sheet {sheet_idx+1}: {ws.title}")
    print(f"Rows: {ws.max_row}")
    
    # Collect player IDs from column A (row 4 onwards = data rows, rows 1-3 are headers)
    # Sheet 1: player_id in col A
    # Sheets 2-4: login_name (phone) in col A, need to resolve to player_id
    
    player_ids = []
    player_rows = {}  # player_id -> row number
    
    col_a_values = []
    for row in range(4, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            col_a_values.append((row, val))
    
    print(f"  Found {len(col_a_values)} entries")
    if not col_a_values:
        continue
    
    # All sheets use login_name in col A (email/phone/etc)
    # We need to resolve login_name -> player_id via sys_player
    login_names = [str(int(v)) if isinstance(v, (float, int)) else str(v) for _, v in col_a_values]
    
    all_mappings = {}
    chunk_size = 500
    resolve_ok = True
    for i in range(0, len(login_names), chunk_size):
        chunk = login_names[i:i+chunk_size]
        placeholders = ','.join(['%s'] * len(chunk))
        try:
            rows = q(f"SELECT player_id, login_name FROM sys_player WHERE login_name IN ({placeholders})", chunk)
            for r in rows:
                all_mappings[str(r['login_name'])] = r['player_id']
            if i == 0:
                print(f"  Resolved {len(rows)} in first chunk")
        except Exception as e:
            print(f"  Error resolving login_names: {e}")
            resolve_ok = False
            break
    
    if resolve_ok and all_mappings:
        for row, val in col_a_values:
            ln = str(int(val)) if isinstance(val, (float, int)) else str(val)
            pid = all_mappings.get(ln)
            if pid:
                player_rows[pid] = row
                player_ids.append(pid)
        print(f"  Mapped {len(player_ids)}/{len(col_a_values)} via login_name")
    else:
        # Fallback: try col A as player_id directly
        print("  Trying col A as player_id...")
        for row, val in col_a_values:
            try:
                pid = int(val)
                player_rows[pid] = row
                player_ids.append(pid)
            except (ValueError, TypeError):
                pass
    
    print(f"  Resolved {len(player_ids)} player_ids")
    
    if not player_ids:
        print("  SKIPPING - no player_ids resolved")
        continue
    
    # Query player info + daily deposits in chunks
    all_daily = {}  # player_id -> {date -> deposit}
    all_info = {}   # player_id -> {vip, reg_time, total_deposit, total_bet, total_win}
    
    chunk_size = 2000
    for i in range(0, len(player_ids), chunk_size):
        chunk = player_ids[i:i+chunk_size]
        id_str = ','.join(str(x) for x in chunk)
        
        # Daily deposits (1/24 - 2/8)
        rows = q(f"""
            SELECT player_id, statistics_day, recharge_amount
            FROM player_statistics_day
            WHERE player_id IN ({id_str})
            AND statistics_day BETWEEN '2026-01-24' AND '2026-02-08'
        """)
        for r in rows:
            pid = r['player_id']
            if pid not in all_daily:
                all_daily[pid] = {}
            all_daily[pid][r['statistics_day']] = float(r['recharge_amount'] or 0)
        
        # Historical totals
        rows = q(f"""
            SELECT player_id,
                SUM(recharge_amount) as total_dep,
                SUM(bet_amount) as total_bet,
                SUM(profit_amount) as total_win
            FROM player_statistics_day
            WHERE player_id IN ({id_str})
            GROUP BY player_id
        """)
        for r in rows:
            all_info[r['player_id']] = {
                'total_deposit': float(r['total_dep'] or 0),
                'total_bet': float(r['total_bet'] or 0),
                'total_win': float(r['total_win'] or 0),
            }
        
        print(f"  Chunk {i//chunk_size + 1}: fetched {len(rows)} player summaries")
    
    # Try to get VIP level and reg time from vip_oper_log or first_deposit_record
    for i in range(0, len(player_ids), chunk_size):
        chunk = player_ids[i:i+chunk_size]
        id_str = ','.join(str(x) for x in chunk)
        
        # VIP level (latest)
        try:
            rows = q(f"""
                SELECT player_id, MAX(after_level) as vip_level
                FROM vip_oper_log
                WHERE player_id IN ({id_str})
                GROUP BY player_id
            """)
            for r in rows:
                if r['player_id'] in all_info:
                    all_info[r['player_id']]['vip'] = r['vip_level']
                else:
                    all_info[r['player_id']] = {'vip': r['vip_level']}
        except Exception as e:
            print(f"  VIP query error: {e}")
        
        # Registration time from first_deposit_record
        try:
            rows = q(f"""
                SELECT player_id, MIN(create_time) as reg_time
                FROM first_deposit_record
                WHERE player_id IN ({id_str})
                GROUP BY player_id
            """)
            for r in rows:
                if r['player_id'] in all_info:
                    all_info[r['player_id']]['reg_time'] = r['reg_time']
                else:
                    all_info[r['player_id']] = {'reg_time': r['reg_time']}
        except Exception as e:
            print(f"  Reg time query error: {e}")
    
    # Fill the worksheet
    filled = 0
    for pid, row in player_rows.items():
        info = all_info.get(pid, {})
        daily = all_daily.get(pid, {})
        
        # Col B: 会员ID (if not already filled)
        if not ws.cell(row=row, column=2).value:
            ws.cell(row=row, column=2, value=pid)
        
        # Col C: VIP等级
        vip = info.get('vip')
        if vip is not None:
            ws.cell(row=row, column=3, value=int(vip))
        
        # Col D: 账号注册时间
        reg = info.get('reg_time')
        if reg:
            ws.cell(row=row, column=4, value=reg)
        
        # Col E: 历史累计存款
        ws.cell(row=row, column=5, value=info.get('total_deposit', 0))
        
        # Col F: 历史累计投注
        ws.cell(row=row, column=6, value=info.get('total_bet', 0))
        
        # Col G: 历史总输赢
        ws.cell(row=row, column=7, value=info.get('total_win', 0))
        
        # Daily deposit columns
        for col_idx, d in date_cols_corrected.items():
            dep = daily.get(d, 0)
            ws.cell(row=row, column=col_idx+1, value=dep)  # +1 because openpyxl is 1-indexed
        
        filled += 1
    
    print(f"  Filled {filled} rows")

# Save
out_path = '/Users/sulaxd/clawd/output/recall_activity_filled.xlsx'
wb.save(out_path)
print(f"\n✅ Saved to {out_path}")
conn.close()
